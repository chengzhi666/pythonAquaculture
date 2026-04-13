"""
Summarize real manual-evaluation scores from sampled workbooks.

支持:
  - 单评审员: --input reviewer.xlsx
  - 双评审员: --input reviewer_a.xlsx --input2 reviewer_b.xlsx
    自动计算 Cohen's κ 评审者间一致性（按维度 + 综合）

Usage:
    # 单评审员
    python summarize_manual_eval.py --input results/.../manual_eval_100_samples.xlsx

    # 双评审员（含 Cohen's κ）
    python summarize_manual_eval.py \\
        --input results/.../reviewer_a.xlsx \\
        --input2 results/.../reviewer_b.xlsx
"""

from __future__ import annotations

import argparse
import json
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_SCORE_FIELDS = ("fact_correctness", "logic_coherence", "completeness")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Summarize manual evaluation workbook scores.")
    parser.add_argument(
        "--input",
        required=True,
        help="Path to the XLSX workbook produced by manual_eval_sampler.py (reviewer A).",
    )
    parser.add_argument(
        "--input2",
        default="",
        help="Optional second XLSX workbook (reviewer B) to compute inter-rater agreement.",
    )
    parser.add_argument(
        "--output-json",
        default="",
        help="Optional output JSON path. Defaults to <input>_summary.json",
    )
    return parser.parse_args()


def to_float(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _read_workbook(input_path: Path) -> list[dict]:
    """读取工作簿并返回有完整打分的行列表。"""
    workbook = load_workbook(input_path)
    sheet = workbook.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {str(name): idx for idx, name in enumerate(headers)}

    missing_headers = [field for field in REQUIRED_SCORE_FIELDS if field not in header_index]
    if missing_headers:
        raise SystemExit(f"Workbook {input_path} is missing required columns: {missing_headers}")

    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        payload = {str(headers[idx]): row[idx] for idx in range(len(headers))}
        fact = to_float(payload.get("fact_correctness"))
        logic = to_float(payload.get("logic_coherence"))
        completeness = to_float(payload.get("completeness"))
        if None in (fact, logic, completeness):
            continue
        overall = round((fact + logic + completeness) / 3, 4)
        rows.append(
            {
                "sample_no": payload.get("sample_no"),
                "template_type": payload.get("template_type", "unknown"),
                "fact_correctness": fact,
                "logic_coherence": logic,
                "completeness": completeness,
                "overall_score": overall,
            }
        )
    return rows


def _compute_avg(rows: list[dict], field: str) -> float:
    if not rows:
        return 0.0
    return round(sum(r[field] for r in rows) / len(rows), 4)


def _per_template_stats(rows: list[dict]) -> dict[str, dict]:
    """按 template_type 分组统计各维度均分。"""
    grouped: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        grouped[r.get("template_type", "unknown")].append(r)

    stats: dict[str, dict] = {}
    for ttype, group_rows in sorted(grouped.items()):
        stats[ttype] = {
            "sample_count": len(group_rows),
            "fact_correctness_avg": _compute_avg(group_rows, "fact_correctness"),
            "logic_coherence_avg": _compute_avg(group_rows, "logic_coherence"),
            "completeness_avg": _compute_avg(group_rows, "completeness"),
            "overall_avg": _compute_avg(group_rows, "overall_score"),
        }
    return stats


# ── Cohen's κ 计算（不依赖 sklearn，纯 Python 实现） ─────────────


def _cohens_kappa(ratings_a: list[int], ratings_b: list[int]) -> float:
    """计算 Cohen's κ 系数。

    将连续评分四舍五入为整数离散化后计算。
    当两位评审员的评分无变异（全部相同）时返回 1.0。
    """
    if len(ratings_a) != len(ratings_b) or not ratings_a:
        return 0.0

    # 离散化为整数
    a = [round(x) for x in ratings_a]
    b = [round(x) for x in ratings_b]

    categories = sorted(set(a) | set(b))
    if len(categories) <= 1:
        # 两位评审员全部给了相同分数
        return 1.0

    n = len(a)
    cat_idx = {c: i for i, c in enumerate(categories)}
    k = len(categories)

    # 混淆矩阵
    matrix = [[0] * k for _ in range(k)]
    for ai, bi in zip(a, b):
        matrix[cat_idx[ai]][cat_idx[bi]] += 1

    # 观察一致率
    po = sum(matrix[i][i] for i in range(k)) / n

    # 期望一致率
    pe = 0.0
    for i in range(k):
        row_sum = sum(matrix[i][j] for j in range(k))
        col_sum = sum(matrix[j][i] for j in range(k))
        pe += (row_sum / n) * (col_sum / n)

    if abs(1.0 - pe) < 1e-12:
        return 1.0

    kappa = (po - pe) / (1.0 - pe)
    return round(kappa, 4)


def _compute_agreement(rows_a: list[dict], rows_b: list[dict]) -> dict:
    """对齐两位评审员的打分（按 sample_no 匹配），计算各维度 Cohen's κ。"""
    map_b = {r["sample_no"]: r for r in rows_b}

    paired_a: list[dict] = []
    paired_b: list[dict] = []
    for r in rows_a:
        key = r["sample_no"]
        if key in map_b:
            paired_a.append(r)
            paired_b.append(map_b[key])

    if not paired_a:
        return {"error": "No matching sample_no found between the two workbooks."}

    agreement: dict = {"paired_samples": len(paired_a)}
    for field in (*REQUIRED_SCORE_FIELDS, "overall_score"):
        a_scores = [r[field] for r in paired_a]
        b_scores = [r[field] for r in paired_b]
        kappa = _cohens_kappa(a_scores, b_scores)
        agreement[f"{field}_kappa"] = kappa

    return agreement


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)

    rows_a = _read_workbook(input_path)
    if not rows_a:
        raise SystemExit("No completed score rows found in workbook.")

    summary: dict = {
        "input_path": str(input_path),
        "completed_rows": len(rows_a),
        "fact_correctness_avg": _compute_avg(rows_a, "fact_correctness"),
        "logic_coherence_avg": _compute_avg(rows_a, "logic_coherence"),
        "completeness_avg": _compute_avg(rows_a, "completeness"),
        "overall_avg": _compute_avg(rows_a, "overall_score"),
        "per_template": _per_template_stats(rows_a),
    }

    # ── 双评审员 Cohen's κ ────────────────────────
    if args.input2:
        input2_path = Path(args.input2)
        rows_b = _read_workbook(input2_path)
        if rows_b:
            agreement = _compute_agreement(rows_a, rows_b)
            summary["inter_rater_agreement"] = agreement
            summary["reviewer_b_path"] = str(input2_path)
            summary["reviewer_b_completed_rows"] = len(rows_b)

            # 同时计算 B 的均分
            summary["reviewer_b_overall_avg"] = _compute_avg(rows_b, "overall_score")

            print(f"Reviewer A rows: {len(rows_a)}")
            print(f"Reviewer B rows: {len(rows_b)}")
            print(f"Paired samples: {agreement.get('paired_samples', 0)}")
            for field in (*REQUIRED_SCORE_FIELDS, "overall_score"):
                kappa_key = f"{field}_kappa"
                print(f"  {field} Cohen's κ = {agreement.get(kappa_key, 'N/A')}")
        else:
            print(f"Warning: Reviewer B workbook has no completed rows: {input2_path}")

    output_json = Path(args.output_json) if args.output_json else input_path.with_name(f"{input_path.stem}_summary.json")
    output_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"\nCompleted rows: {summary['completed_rows']}")
    print(f"Overall average: {summary['overall_avg']}")
    if summary.get("per_template"):
        print("Per-template breakdown:")
        for ttype, stats in summary["per_template"].items():
            print(f"  {ttype}: n={stats['sample_count']}, avg={stats['overall_avg']}")
    print(f"Summary JSON: {output_json}")


if __name__ == "__main__":
    main()
